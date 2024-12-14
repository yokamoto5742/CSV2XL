import os
from os import startfile
from pathlib import Path
import polars as pl
from openpyxl import load_workbook
import datetime
from config_manager import ConfigManager

VERSION = "0.0.2"
LAST_UPDATED = "2024/12/12"

def read_csv_with_encoding(file_path):
    encodings = ['shift-jis', 'utf-8']

    for encoding in encodings:
        try:
            schema = {
                "患者ID": pl.Int64,
            }
            df = pl.read_csv(
                file_path,
                encoding=encoding,
                separator=',',
                skip_rows=3,  # 最初の3行をスキップ
                has_header=True,  # 4行目をヘッダーとして使用
                infer_schema_length=0,
                schema_overrides=schema
            )

            if len(df.columns) > 1:
                print(f"エンコーディング {encoding} で正常に読み込みました")
                print(f"列数: {len(df.columns)}")
                print(f"行数: {len(df)}")
                print(f"列名: {df.columns}")
                return df
        except Exception as e:
            print(f"{encoding}での読み込み試行中にエラー: {str(e)}")
            continue

    raise Exception("CSVファイルの読み込みに失敗しました")

def process_csv_data(df):
    try:
        print("処理前の列名:", df.columns)
        # 列名を一意にする
        original_columns = df.columns
        unique_columns = []
        for i, col in enumerate(original_columns):
            unique_columns.append(f"col_{i}_{col}")
        df = df.select([
            pl.col(old_name).alias(new_name)
            for old_name, new_name in zip(original_columns, unique_columns)
        ])

        # K列とI列を削除 (インデックスベースで削除)
        columns_to_keep = [i for i in range(len(df.columns)) if i not in [8, 10]]
        df = df.select([df.columns[i] for i in columns_to_keep])

        # A列からC列を削除
        df = df.select(df.columns[3:])

        config = ConfigManager()
        exclude_docs = config.get_exclude_docs()

        if exclude_docs:
            filter_conditions = [~pl.col(df.columns[3]).str.contains(doc) for doc in exclude_docs]
            combined_filter = filter_conditions[0]
            for condition in filter_conditions[1:]:
                combined_filter = combined_filter & condition
            df = df.filter(combined_filter)

        print("処理後の列名:", df.columns)
        return df

    except Exception as e:
        print(f"データ処理中にエラーが発生しました: {str(e)}")
        raise


def transfer_csv_to_excel():
    try:
        config = ConfigManager()
        downloads_path = config.get_downloads_path()
        excel_path = config.get_excel_path()

        # CSVファイルを検索（最新のCSVファイルを使用）
        csv_files = [f for f in os.listdir(downloads_path) if f.endswith('.csv')]
        if not csv_files:
            print("ダウンロードフォルダにCSVファイルが見つかりません。")
            return

        latest_csv = max([os.path.join(downloads_path, f) for f in csv_files],
                         key=os.path.getmtime)

        print(f"処理するCSVファイル: {latest_csv}")

        df = read_csv_with_encoding(latest_csv)

        try:
            date_col = df.columns[3]
            df = df.with_columns([
                pl.col(date_col).str.strptime(pl.Date, format="%Y%m%d")
                .alias(date_col)
            ])
        except Exception as e:
            print(f"日付変換中にエラーが発生しましたが、処理を継続します: {str(e)}")

        df = process_csv_data(df)

        if not os.path.exists(excel_path):
            print(f"Excelファイルが見つかりません: {excel_path}")
            return

        wb = load_workbook(excel_path)
        ws = wb.active

        # 最終行を取得
        last_row = ws.max_row

        temp_df = df.select([
            pl.col('*').cast(pl.String)
        ])
        data_to_write = temp_df.to_numpy().tolist()

        for i, row in enumerate(data_to_write):
            for j, value in enumerate(row):
                cell = ws.cell(row=last_row + 1 + i, column=j + 1)

                if j == 0:  # 預り日（日付列）
                    try:
                        # 文字列から日付に戻す
                        date_value = datetime.datetime.strptime(value, '%Y-%m-%d')
                        cell.value = date_value
                        cell.number_format = 'yyyy/mm/dd'
                    except:
                        cell.value = value
                elif j == 1:  # 患者ID列
                    try:
                        cell.value = int(value.replace(',', ''))
                        cell.number_format = '0'
                    except:
                        cell.value = value
                else:
                    cell.value = value if value is not None else ""

        wb.save(excel_path)
        print("データの転記が完了しました。")

        startfile(excel_path)

    except Exception as e:
        print(f"エラーが発生しました: {str(e)}")
        import traceback
        print("詳細なエラー情報:")
        print(traceback.format_exc())

if __name__ == "__main__":
    transfer_csv_to_excel()
