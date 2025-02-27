# excel_processor.py - Excelファイルの処理に関する機能
import os
from pathlib import Path
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time
import pyautogui
import win32com.client
import win32gui
from PyQt6.QtWidgets import QMessageBox
from file_manager import FileManager


class ExcelProcessor:
    """Excelファイルの処理を行うクラス"""
    
    def process_excel_file(self, excel_path, df):
        """Excelファイルを開いてデータを書き込む"""
        if not os.path.exists(excel_path) or not excel_path.endswith('.xlsm'):
            QMessageBox.critical(None, "エラー", "マクロ付きExcelファイル(.xlsm)が見つかりません。")
            return False

        try:
            wb = load_workbook(filename=excel_path, read_only=False, keep_vba=True)
        except PermissionError:
            QMessageBox.critical(None, "エラー", 
                                "Excelファイルが別のプロセスで開かれています。\nExcelファイルを閉じてから再度実行してください。")
            return False
            
        ws = wb.active

        # 実際のデータが存在する最終行を取得
        last_row = self._get_last_row(ws)

        # 既存データを保持するセットを作成
        existing_data = self._get_existing_data(ws, last_row)

        # CSVデータを文字列に変換
        from csv_processor import CSVProcessor
        csv_processor = CSVProcessor()
        data_to_write = csv_processor.convert_to_string_data(df)

        # 重複していないデータのみを抽出
        unique_data = self._filter_unique_data(data_to_write, existing_data)

        # データをExcelに書き込む
        self._write_data_to_excel(ws, unique_data, last_row)
        
        # セルの書式を適用
        self._apply_cell_formats(ws, last_row + 1)

        try:
            wb.save(excel_path)
            wb.close()
            file_manager = FileManager(None)
            file_manager.backup_excel_file(excel_path)
            return True
        except PermissionError:
            QMessageBox.critical(None, "エラー",
                                "Excelファイルが別のプロセスで開かれているため、保存できません。\nExcelファイルを閉じてから再度実行してください。")
            if 'wb' in locals():
                wb.close()
            return False

    def _get_last_row(self, worksheet):
        """ワークシートの最終行を取得"""
        last_row = 0
        for row in worksheet.iter_rows():
            if all(cell.value is None for cell in row):
                break
            last_row += 1
        return last_row

    def _apply_cell_formats(self, worksheet, start_row):
        """セルの書式設定を適用"""
        last_row = self._get_last_row(worksheet)

        for row in range(start_row, last_row + 1):
            for col in range(1, 7):  # A=1, F=6
                cell = worksheet.cell(row=row, column=col)
                
                # 縦位置を中央に設定
                cell.alignment = Alignment(vertical='center')
                
                # 横位置を設定
                if col in [1, 2, 5, 6]:  # A,B列とE,F列
                    cell.alignment = Alignment(vertical='center', horizontal='center')
                elif col in [3, 4]:  # C列とD列
                    cell.alignment = Alignment(vertical='center', horizontal='left', shrink_to_fit=True)

    def _get_existing_data(self, worksheet, last_row):
        """既存データのセットを作成"""
        existing_data = set()
        for row in range(2, last_row + 1):  # ヘッダー行をスキップ
            # 日付をYYYYMMDD形式の文字列として取得
            date_value = worksheet.cell(row=row, column=1).value
            if isinstance(date_value, datetime.datetime):
                date_str = date_value.strftime('%Y%m%d')
            else:
                date_str = str(date_value or '')

            # A列からF列までの値を取得
            row_data = (
                date_str,
                str(worksheet.cell(row=row, column=2).value or ''),
                str(worksheet.cell(row=row, column=3).value or ''),
                str(worksheet.cell(row=row, column=4).value or ''),
                str(worksheet.cell(row=row, column=5).value or ''),
                str(worksheet.cell(row=row, column=6).value or '')
            )
            existing_data.add(row_data)
        
        return existing_data

    def _filter_unique_data(self, data_to_write, existing_data):
        """重複していないデータのみを抽出"""
        unique_data = []
        for row in data_to_write:
            # CSVの日付を8桁の数値文字列に変換
            csv_date = row[0]
            if isinstance(csv_date, str):
                try:
                    # YYYY-MM-DD形式をYYYYMMDD形式に変換
                    date_obj = datetime.datetime.strptime(csv_date, '%Y-%m-%d')
                    date_str = date_obj.strftime('%Y%m%d')
                except:
                    date_str = csv_date
            else:
                date_str = str(csv_date)

            # 比較用のタプルを作成
            row_data = (
                date_str,
                str(row[1] or ''),
                str(row[2] or ''),
                str(row[3] or ''),
                str(row[4] or ''),
                str(row[5] or '')
            )

            if row_data not in existing_data:
                unique_data.append(row)
        
        return unique_data

    def _write_data_to_excel(self, worksheet, unique_data, last_row):
        """Excelにデータを書き込む"""
        for i, row in enumerate(unique_data):
            for j, value in enumerate(row):
                cell = worksheet.cell(row=last_row + 1 + i, column=j + 1)

                if j == 0:  # 日付列
                    try:
                        date_value = datetime.datetime.strptime(value, '%Y-%m-%d')
                        cell.value = date_value
                        cell.number_format = 'yyyy/mm/dd'
                    except:
                        cell.value = value
                elif j == 1:  # 患者ID列
                    try:
                        cell.value = int(value.replace(',', ''))
                        cell.number_format = '0'
                    except:
                        cell.value = value
                else:
                    cell.value = value if value is not None else ""

    def open_and_process_excel(self, excel_path, config):
        """Excelファイルを開いて共有ボタンをクリック"""
        excel_path_str = str(Path(excel_path).resolve())
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        self._bring_excel_to_front()
        workbook = excel.Workbooks.Open(excel_path_str)
        excel.WindowState = -4137  # xlMaximized
        workbook.Windows(1).Activate()

        try:
            worksheet = workbook.ActiveSheet
            self._sort_excel_data(worksheet)

            last_row = worksheet.Cells(worksheet.Rows.Count, "A").End(-4162).Row
            worksheet.Cells(last_row, 1).Select()

            wait_time = config.get_share_button_wait_time()
            time.sleep(wait_time)
            share_x, share_y = config.get_share_button_position()
            pyautogui.click(share_x, share_y)

        except Exception as e:
            print(f"共有ボタンのクリックに失敗しました: {str(e)}")
        finally:
            # 操作が終わったらExcelは開いたままにする
            pyautogui.hotkey('win', 'down')  # ウィンドウを最小化
            excel = None

    def _bring_excel_to_front(self):
        """Excelを最前面に表示"""
        for _ in range(2):
            hwnd = win32gui.FindWindow("XLMAIN", None)
            if hwnd:
                win32gui.SetForegroundWindow(hwnd)
                return True
            time.sleep(0.1)
        return False

    def _sort_excel_data(self, worksheet):
        """Excelデータを指定列でソート"""
        try:
            # 最終行を取得 (-4162はxlUpに相当)
            last_row = worksheet.Cells(worksheet.Rows.Count, "A").End(-4162).Row

            # ソート対象の範囲を設定
            sort_range = worksheet.Range(f"A2:I{last_row}")

            # ソートの実行
            sort_range.Sort(
                Key1=worksheet.Range("A2"),  # A列（預り日）
                Order1=1,                    # 1=昇順
                Key2=worksheet.Range("E2"),  # E列（診療科）
                Order2=1,                    # 1=昇順
                Key3=worksheet.Range("B2"),  # B列（患者ID）
                Order3=1,                    # 1=昇順
                Header=1,                    # 1=ヘッダーあり
                OrderCustom=1,
                MatchCase=False,
                Orientation=1
            )

            return last_row

        except Exception as e:
            print(f"ソート中にエラーが発生しました: {str(e)}")
            raise
