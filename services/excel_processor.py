import datetime
import time
from pathlib import Path

import polars as pl
import pyautogui
import win32com.client
import win32gui
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from PyQt6.QtWidgets import QMessageBox

from utils.config_manager import ConfigManager


def get_last_row(worksheet):
    """ワークシートの最後のデータ行番号を取得

    Args:
        worksheet: openpyxlのワークシートオブジェクト

    Returns:
        最後のデータが存在する行番号
    """
    last_row = 0
    for row in worksheet.iter_rows():
        if all(cell.value is None for cell in row):
            break
        last_row += 1
    return last_row


def apply_cell_formats(worksheet, start_row):
    """指定された行からセルのフォーマットと配置を適用

    Args:
        worksheet: openpyxlのワークシートオブジェクト
        start_row: フォーマット開始行番号
    """
    last_row = get_last_row(worksheet)

    # A～F列に対して、垂直中央配置を設定し、特定列の水平配置を調整
    for row in range(start_row, last_row + 1):
        for col in range(1, 7):
            cell = worksheet.cell(row=row, column=col)

            cell.alignment = Alignment(vertical='center')

            if col in [1, 2, 5, 6]:  # A,B列とE,F列
                cell.alignment = Alignment(horizontal='center')

            elif col in [3, 4]:  # C列とD列
                cell.alignment = Alignment(horizontal='left', shrink_to_fit=True)


def sort_excel_data(worksheet):
    """Excelデータを預り日、診療科、患者IDでソート

    Args:
        worksheet: win32comのワークシートオブジェクト

    Returns:
        ソート後の最終行番号
    """
    try:
        last_row = worksheet.Cells(worksheet.Rows.Count, "A").End(-4162).Row

        # ソートの範囲を設定
        sort_range = worksheet.Range(f"A2:I{last_row}")

        sort_range.Sort(
            Key1=worksheet.Range("A2"),  # A列（預り日）
            Order1=1,  # 1=昇順
            Key2=worksheet.Range("E2"),  # E列（診療科）
            Order2=1,  # 1=昇順
            Key3=worksheet.Range("B2"),  # B列（患者ID）
            Order3=1,  # 1=昇順
            Header=1,  # 1=ヘッダーあり
            OrderCustom=1,
            MatchCase=False,
            Orientation=1)

        return last_row

    except Exception as e:
        print(f"ソート中にエラーが発生しました: {str(e)}")
        raise


def bring_excel_to_front():
    """Excelウィンドウを最前面に表示

    Returns:
        成功時はTrue、失敗時はFalse
    """
    for _ in range(2):
        hwnd = win32gui.FindWindow("XLMAIN", None)
        if hwnd:
            win32gui.SetForegroundWindow(hwnd)
            return True
        time.sleep(0.1)
    return False


def write_data_to_excel(excel_path, df):
    """DataFrameのデータをExcelファイルに重複排除して書き込み

    既存データを確認して重複していないデータのみを追加。日付と患者IDの形式変換も実施

    Args:
        excel_path: Excelファイルのパス
        df: 書き込むpolarsのDataFrame

    Returns:
        成功時はTrue、失敗時はFalse
    """
    if not Path(excel_path).exists() or not excel_path.endswith('.xlsm'):
        print(f"Excelファイルが見つかりません: {excel_path}")
        return False

    try:
        wb = load_workbook(filename=excel_path, keep_vba=True)
    except PermissionError:
        QMessageBox.critical(None,
                             "エラー",
                             "Excelファイルが別のプロセスで開かれています。\nファイルを閉じてから再度実行してください。"
                             )
        return False

    ws = wb.active

    last_row = get_last_row(ws)

    # 既存データのセットを構築して重複チェック用のキーを作成（A～F列の値で識別）
    existing_data = set()
    for row in range(2, last_row + 1):
        # 日付をYYYYMMDD形式の文字列として取得
        cell1 = ws.cell(row=row, column=1)
        date_value = cell1.value if cell1 else None
        if isinstance(date_value, datetime.datetime):
            date_str = date_value.strftime('%Y%m%d')
        else:
            date_str = str(date_value or '')

        # A列からF列までの値を取得（日付は数値形式で保持）
        cell2 = ws.cell(row=row, column=2)
        cell3 = ws.cell(row=row, column=3)
        cell4 = ws.cell(row=row, column=4)
        cell5 = ws.cell(row=row, column=5)
        cell6 = ws.cell(row=row, column=6)
        row_data = (
            date_str,  # 日付を8桁の数値文字列として保持
            str(cell2.value or '') if cell2 else '',
            str(cell3.value or '') if cell3 else '',
            str(cell4.value or '') if cell4 else '',
            str(cell5.value or '') if cell5 else '',
            str(cell6.value or '') if cell6 else ''
        )
        existing_data.add(row_data)

    # DataFrameのすべての値を文字列に変換
    temp_df = df.select([
        pl.col('*').cast(pl.String)
    ])
    data_to_write = temp_df.to_numpy().tolist()

    # 既存データセットに存在しないデータのみを抽出
    unique_data = []
    for row in data_to_write:
        # 日付形式をYYYYMMDD形式に統一して比較
        csv_date = row[0]
        if isinstance(csv_date, str):
            try:
                # YYYY-MM-DD形式をYYYYMMDD形式に変換
                date_obj = datetime.datetime.strptime(csv_date, '%Y-%m-%d')
                date_str = date_obj.strftime('%Y%m%d')
            except ValueError:
                date_str = csv_date
        else:
            date_str = str(csv_date)

        # 比較用のタプルを作成
        row_data = (
            date_str,
            str(row[1] or ''),
            str(row[2] or ''),
            str(row[3] or ''),
            str(row[4] or ''),
            str(row[5] or '')
        )

        if row_data not in existing_data:
            unique_data.append(row)

    # 新規データを行ごとにセルに書き込み、必要に応じて型変換を実施
    for i, row in enumerate(unique_data):
        for j, value in enumerate(row):
            cell = ws.cell(row=last_row + 1 + i, column=j + 1)

            if j == 0:  # 日付列を日付型に変換
                try:
                    date_value = datetime.datetime.strptime(value, '%Y-%m-%d')
                    cell.value = date_value
                    cell.number_format = 'yyyy/mm/dd'
                except ValueError:
                    cell.value = value
            elif j == 1:  # 患者ID列
                try:
                    cell.value = int(value.replace(',', ''))
                    cell.number_format = '0'
                except ValueError:
                    cell.value = value
            else:
                cell.value = value if value is not None else ""

    apply_cell_formats(ws, last_row + 1)

    try:
        wb.save(excel_path)
        wb.close()
        return True
    except PermissionError:
        QMessageBox.critical(None,
                             "エラー",
                             "Excelファイルが別のプロセスで開かれているため、保存できません。\nファイルを閉じてから再度実行してください。"
                             )
        if 'wb' in locals():
            wb.close()
        return False


def clear_all_filters(worksheet, workbook):
    """ワークシートのフィルタをすべてクリアして解除"""
    try:
        is_shared = False
        try:
            # 共有ブックかどうかを確認
            is_shared = workbook.MultiUserEditing
        except Exception:
            pass

        if not worksheet.AutoFilterMode:
            print("オートフィルタは設定されていません")
            return

        # フィルタ条件が適用されていればすべてのデータを表示
        try:
            if worksheet.FilterMode:
                worksheet.ShowAllData()
                print("フィルタ条件をクリアしました")
        except Exception as e:
            print(f"ShowAllData実行中にエラー: {str(e)}")

        # 念のため各列のフィルタを個別にクリアする
        try:
            if worksheet.AutoFilterMode:
                auto_filter = worksheet.AutoFilter
                if auto_filter is not None:
                    filters = auto_filter.Filters
                    for i in range(1, filters.Count + 1):
                        try:
                            if filters.Item(i).On:
                                # 各列のフィルタ条件をクリア
                                auto_filter.Range.AutoFilter(Field=i)
                        except Exception:
                            pass
                    print("個別のフィルタ条件をクリアしました")
        except Exception as e:
            print(f"個別フィルタクリア中にエラー: {str(e)}")

        # 共有ブックでない場合のみオートフィルタ自体を解除
        if not is_shared:
            try:
                worksheet.AutoFilterMode = False
                print("オートフィルタを解除しました")
            except Exception as e:
                print(f"AutoFilterMode解除中にエラー（共有ブックの可能性）: {str(e)}")
        else:
            print("共有ブックのため、オートフィルタの解除はスキップしました（フィルタ条件はクリア済み）")

    except Exception as e:
        print(f"フィルタ解除中にエラーが発生しました: {str(e)}")


def open_and_sort_excel(excel_path):
    """Excelファイルを開いてデータをソート、共有ボタンをクリック

    Args:
        excel_path: Excelファイルのパス
    """
    excel_path_obj = Path(excel_path)

    if not excel_path_obj.exists():
        QMessageBox.critical(None, "エラー", f"Excelファイルが見つかりません: {excel_path}")
        return

    excel_path_str = str(excel_path_obj.resolve())
    excel = None
    workbook = None

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        bring_excel_to_front()

        # Excelファイルを開く
        workbook = excel.Workbooks.Open(excel_path_str)

        if workbook is None:
            QMessageBox.critical(None, "エラー", "Excelファイルを開くことができませんでした。")
            return

        excel.WindowState = -4137  # xlMaximized
        workbook.Windows(1).Activate()

        worksheet = workbook.ActiveSheet

        # フィルタがかかっていればクリア
        clear_all_filters(worksheet, workbook)

        # データをソート
        sort_excel_data(worksheet)

        # 最終行にカーソルを移動
        last_row = worksheet.Cells(worksheet.Rows.Count, "A").End(-4162).Row
        worksheet.Cells(last_row, 1).Select()

        # 設定に従って共有ボタンをクリック
        config = ConfigManager()
        wait_time = config.get_share_button_wait_time()
        time.sleep(wait_time)
        share_x, share_y = config.get_share_button_position()
        pyautogui.click(share_x, share_y)

    except Exception as e:
        error_msg = f"Excelファイルの処理中にエラーが発生しました: {str(e)}"
        print(error_msg)
        QMessageBox.critical(None, "エラー", error_msg)
    finally:
        # Excelは開いたままにするがエラー処理は行う
        pyautogui.hotkey('win', 'down')  # ウィンドウを最小化
