import os
from os import startfile
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


def process_csv_data(df):
    """CSVデータの加工処理を行う"""
    # 最初の4行を削除
    df = df.iloc[4:]

    # インデックスをリセット
    df = df.reset_index(drop=True)

    # K列を削除
    df = df.drop(df.columns[10], axis=1)

    # I列を削除
    df = df.drop(df.columns[8], axis=1)

    # A列からC列を削除
    columns_to_drop = df.columns[:3]
    df = df.drop(columns=columns_to_drop)

    return df


def transfer_csv_to_excel():
    try:
        # ダウンロードフォルダのパスを取得
        downloads_path = str(Path.home() / "Downloads")

        # 対象のExcelファイルのパス
        excel_path = r"C:\Shinseikai\CSV2XL\医療文書担当一覧.xlsx"

        # CSVファイルを検索（最新のCSVファイルを使用）
        csv_files = [f for f in os.listdir(downloads_path) if f.endswith('.csv')]
        if not csv_files:
            print("ダウンロードフォルダにCSVファイルが見つかりません。")
            return

        latest_csv = max([os.path.join(downloads_path, f) for f in csv_files],
                         key=os.path.getmtime)

        # CSVファイルを読み込む
        try:
            df = pd.read_csv(latest_csv, encoding='utf-8')
        except UnicodeDecodeError:
            # UTF-8で読み込めない場合はCP932（Shift-JIS）で試行
            df = pd.read_csv(latest_csv, encoding='cp932')

        try:
            df = pd.read_csv(latest_csv,
                             encoding='utf-8',
                             parse_dates=[0],  # 最初の列（A列）を日付として解析
                             dayfirst=False)  # 月/日/年の形式を想定
        except UnicodeDecodeError:
            # UTF-8で読み込めない場合はCP932（Shift-JIS）で試行
            df = pd.read_csv(latest_csv,
                             encoding='cp932',
                             parse_dates=[0],  # 最初の列（A列）を日付として解析
                             dayfirst=False)  # 月/日/年の形式を想定

        # CSVデータの加工処理
        df = process_csv_data(df)

        # Excelファイルを読み込む
        if not os.path.exists(excel_path):
            print(f"Excelファイルが見つかりません: {excel_path}")
            return

        wb = load_workbook(excel_path)
        ws = wb.active

        # 最終行を取得
        last_row = ws.max_row

        # DataFrameのデータをExcelに書き込む
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                # last_row + 1から開始し、データを書き込む
                ws.cell(row=last_row + 1 + i, column=j + 1, value=value)

        # 変更を保存
        wb.save(excel_path)
        print("データの転記が完了しました。")

        startfile(excel_path)

    except Exception as e:
        print(f"エラーが発生しました: {str(e)}")


if __name__ == "__main__":
    transfer_csv_to_excel()