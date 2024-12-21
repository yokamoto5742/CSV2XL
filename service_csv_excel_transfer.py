import os
from os import startfile
from pathlib import Path
import shutil
import polars as pl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import datetime
from config_manager import ConfigManager
from PyQt6.QtWidgets import QMessageBox
from PyQt6.QtCore import QTimer
import time
import pyautogui
import win32com.client

def read_csv_with_encoding(file_path):
    encodings = ['shift-jis', 'utf-8']

    for encoding in encodings:
        try:
            schema = {
                "患者ID": pl.Int64,
            }
            df = pl.read_csv(
                file_path,
                encoding=encoding,
                separator=',',
                skip_rows=3,  # 最初の3行をスキップ
                has_header=True,  # 4行目をヘッダーとして使用
                infer_schema_length=0,
                schema_overrides=schema
            )

            if len(df.columns) > 1:
                print(f"エンコーディング {encoding} で正常に読み込みました")
                print(f"列数: {len(df.columns)}")
                print(f"行数: {len(df)}")
                print(f"列名: {df.columns}")
                return df
        except Exception as e:
            print(f"{encoding}での読み込み試行中にエラー: {str(e)}")
            continue

    raise Exception("CSVファイルの読み込みに失敗しました")

def process_csv_data(df):
    try:
        print("処理前の列名:", df.columns)
        # 列名を一意にする
        original_columns = df.columns
        unique_columns = []
        for i, col in enumerate(original_columns):
            unique_columns.append(f"col_{i}_{col}")
        df = df.select([
            pl.col(old_name).alias(new_name)
            for old_name, new_name in zip(original_columns, unique_columns)
        ])

        # K列とI列を削除 (インデックスベースで削除)
        columns_to_keep = [i for i in range(len(df.columns)) if i not in [8, 10]]
        df = df.select([df.columns[i] for i in columns_to_keep])

        # A列からC列を削除
        df = df.select(df.columns[3:])

        config = ConfigManager()
        exclude_docs = config.get_exclude_docs()
        exclude_doctors = config.get_exclude_doctors()

        if exclude_docs:
            filter_conditions = [~pl.col(df.columns[3]).str.contains(doc) for doc in exclude_docs]
            combined_filter = filter_conditions[0]
            for condition in filter_conditions[1:]:
                combined_filter = combined_filter & condition
            df = df.filter(combined_filter)

        if exclude_doctors:
            doctor_filter_conditions = [~pl.col(df.columns[5]).str.contains(doc) for doc in exclude_doctors]
            doctor_combined_filter = doctor_filter_conditions[0]
            for condition in doctor_filter_conditions[1:]:
                doctor_combined_filter = doctor_combined_filter & condition
            df = df.filter(doctor_combined_filter)

            # D列とF列のスペースと*を除去（4列目と6列目）
            df = df.with_columns([
                pl.col(df.columns[3]).str.replace_all(r'[\s*]', ''),  # D列
                pl.col(df.columns[5]).str.replace_all(r'[\s*]', '')  # F列
            ])

            print("処理後の列名:", df.columns)
            return df

        print("処理後の列名:", df.columns)
        return df

    except Exception as e:
        print(f"データ処理中にエラーが発生しました: {str(e)}")
        raise

def get_last_row(worksheet):
    last_row = 0
    for row in worksheet.iter_rows():
        if all(cell.value is None for cell in row):
            break
        last_row += 1
    return last_row


def apply_cell_formats(worksheet, start_row):
    last_row = get_last_row(worksheet)

    # A列からF列までの範囲を設定
    for row in range(start_row, last_row + 1):
        for col in range(1, 7):  # A=1, F=6
            cell = worksheet.cell(row=row, column=col)

            # 縦位置を中央に設定（A-F列）
            cell.alignment = Alignment(vertical='center')

            # 横位置を設定
            if col in [1, 2, 5, 6]:  # A,B列とE,F列
                cell.alignment = Alignment(horizontal='center')

            elif col in [3, 4]:  # C列とD列
                cell.alignment = Alignment(horizontal='left', shrink_to_fit=True)


def backup_excel_file(excel_path):
    config = ConfigManager()
    backup_dir = Path(config.get_backup_path())

    if not backup_dir.exists():
        backup_dir.mkdir(parents=True)

    backup_path = backup_dir / f"backup_{Path(excel_path).name}"

    try:
        # ファイルをコピー（既存のファイルは上書き）
        shutil.copy2(excel_path, backup_path)
        print(f"バックアップを作成しました: {backup_path}")
    except Exception as e:
        print(f"バックアップ作成中にエラーが発生しました: {str(e)}")
        raise

def cleanup_old_files(processed_dir: Path):
    current_time = datetime.datetime.now()
    for file in processed_dir.glob("*.csv"):
        file_time = datetime.datetime.fromtimestamp(file.stat().st_mtime)
        if (current_time - file_time).days >= 3:
            try:
                file.unlink()
                print(f"古いファイルを削除しました: {file}")
            except Exception as e:
                print(f"ファイル削除中にエラーが発生しました: {file} - {str(e)}")

def process_completed_csv(csv_path: str) -> None:
    try:
        csv_file = Path(csv_path)
        if not csv_file.exists():
            print(f"処理対象のCSVファイルが見つかりません: {csv_path}")
            return

        config = ConfigManager()
        processed_dir = Path(config.get_processed_path())
        processed_dir.mkdir(exist_ok=True, parents=True)

        new_path = processed_dir / csv_file.name
        shutil.move(str(csv_file), str(new_path))
        print(f"CSVファイルを移動しました: {new_path}")
        cleanup_old_files(processed_dir)

    except Exception as e:
        print(f"CSVファイルの処理中にエラーが発生しました: {str(e)}")
        raise

def transfer_csv_to_excel():
    try:
        config = ConfigManager()
        downloads_path = config.get_downloads_path()
        excel_path = config.get_excel_path()

        # CSVファイルの検索
        csv_files = [f for f in os.listdir(downloads_path)
                     if f.endswith('.csv') and
                     len(f.split('_')) == 2 and
                     (3 <= len(f.split('_')[0]) <= 4) and
                     len(f.split('_')[1].split('.')[0]) == 14]
        if not csv_files:
            QMessageBox.warning(None, "警告", "ダウンロードフォルダにCSVファイルが見つかりません。")
            return

        latest_csv = max([os.path.join(downloads_path, f) for f in csv_files],
                         key=os.path.getmtime)

        print(f"処理するCSVファイル: {latest_csv}")

        # CSVファイルの読み込み
        df = read_csv_with_encoding(latest_csv)

        try:
            date_col = df.columns[3]
            df = df.with_columns([
                pl.col(date_col).str.strptime(pl.Date, format="%Y%m%d")
                .alias(date_col)
            ])
        except Exception as e:
            print(f"日付変換中にエラーが発生しましたが、処理を継続します: {str(e)}")

        df = process_csv_data(df)

        if not os.path.exists(excel_path) or not excel_path.endswith('.xlsm'):
            print(f"マクロ付きExcelファイル(.xlsm)が見つかりません: {excel_path}")
            return

        try:
            wb = load_workbook(filename=excel_path, read_only=False, keep_vba=True)
        except PermissionError:
            QMessageBox.critical(None,
                                 "エラー",
                                 "Excelファイルが別のプロセスで開かれています。\nExcelファイルを閉じてから再度実行してください。"
                                 )
            return
        ws = wb.active

        # 実際のデータが存在する最終行を取得
        last_row = get_last_row(ws)

        # 既存データを保持するセットを作成（A列からF列までの値をキーとして使用）
        existing_data = set()
        for row in range(2, last_row + 1):  # ヘッダー行をスキップ
            # 日付をYYYYMMDD形式の文字列として取得
            date_value = ws.cell(row=row, column=1).value
            if isinstance(date_value, datetime.datetime):
                date_str = date_value.strftime('%Y%m%d')
            else:
                date_str = str(date_value or '')

            # A列からF列までの値を取得（日付は数値形式で保持）
            row_data = (
                date_str,  # 日付を8桁の数値文字列として保持
                str(ws.cell(row=row, column=2).value or ''),
                str(ws.cell(row=row, column=3).value or ''),
                str(ws.cell(row=row, column=4).value or ''),
                str(ws.cell(row=row, column=5).value or ''),
                str(ws.cell(row=row, column=6).value or '')
            )
            existing_data.add(row_data)

        # CSVデータを文字列に変換
        temp_df = df.select([
            pl.col('*').cast(pl.String)
        ])
        data_to_write = temp_df.to_numpy().tolist()

        # 重複していないデータのみを抽出
        unique_data = []
        for row in data_to_write:
            # CSVの日付を8桁の数値文字列に変換
            csv_date = row[0]
            if isinstance(csv_date, str):
                try:
                    # YYYY-MM-DD形式をYYYYMMDD形式に変換
                    date_obj = datetime.datetime.strptime(csv_date, '%Y-%m-%d')
                    date_str = date_obj.strftime('%Y%m%d')
                except:
                    date_str = csv_date
            else:
                date_str = str(csv_date)

            # 比較用のタプルを作成
            row_data = (
                date_str,  # 日付を8桁の数値文字列として保持
                str(row[1] or ''),
                str(row[2] or ''),
                str(row[3] or ''),
                str(row[4] or ''),
                str(row[5] or '')
            )

            if row_data not in existing_data:
                unique_data.append(row)

        # 重複しないデータのみを書き込み
        for i, row in enumerate(unique_data):
            for j, value in enumerate(row):
                cell = ws.cell(row=last_row + 1 + i, column=j + 1)

                if j == 0:  # 日付列
                    try:
                        date_value = datetime.datetime.strptime(value, '%Y-%m-%d')
                        cell.value = date_value
                        cell.number_format = 'yyyy/mm/dd'
                    except:
                        cell.value = value
                elif j == 1:  # 患者ID列
                    try:
                        cell.value = int(value.replace(',', ''))
                        cell.number_format = '0'
                    except:
                        cell.value = value
                else:
                    cell.value = value if value is not None else ""

        apply_cell_formats(ws, last_row + 1)

        try:
            wb.save(excel_path)
            wb.close()
            backup_excel_file(excel_path)
        except PermissionError:
            QMessageBox.critical(None,
                                 "エラー",
                                 "Excelファイルが別のプロセスで開かれているため、保存できません。\nExcelファイルを閉じてから再度実行してください。"
                                 )
            if 'wb' in locals():
                wb.close()
            return

        msg = QMessageBox()
        msg.setWindowTitle("完了")
        msg.setText("CSVファイルの取り込みが完了しました")
        msg.show()
        QTimer.singleShot(3000, msg.close)

        process_completed_csv(latest_csv)

        excel_path_str = str(Path(excel_path).resolve())

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        workbook = excel.Workbooks.Open(excel_path_str)
        excel.WindowState = -4137  # xlMaximized
        workbook.Windows(1).Activate()

        try:
            share_button = pyautogui.locateOnScreen('share_button.png')
            if share_button:
                button_center = pyautogui.center(share_button)
                pyautogui.click(button_center.x, button_center.y)
                worksheet = workbook.ActiveSheet
                last_row = worksheet.Cells(worksheet.Rows.Count, "A").End(-4162).Row  # xlUp = -4162
                worksheet.Cells(last_row, 1).Select()

        except Exception as e:
            print(f"共有ボタンのクリックに失敗しました: {str(e)}")
        finally:
            # 操作が終わったらExcelは開いたままにする
            excel = None

    except Exception as e:
        print(f"エラーが発生しました: {str(e)}")
        import traceback
        print("詳細なエラー情報:")
        print(traceback.format_exc())